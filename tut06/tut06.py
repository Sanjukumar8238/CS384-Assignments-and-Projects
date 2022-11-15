from ctypes import sizeof                                           # Importing libraries
import pandas as pd
import csv
import os
from datetime import datetime, timedelta
import openpyxl
from openpyxl import workbook,load_workbook
import email, smtplib, ssl
from email import encoders
from email.mime.base import MIMEBase
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText

from datetime import datetime
start_time = datetime.now()

def attendance_report():                                            # Function to process input files
    try:
        df=pd.read_csv('input_registered_students.csv')             # Reading input_registered_students file
    except:
        print('Input file error')

    rolls={}                                                        # Mapping each roll no. with an integer
    roll_numbers=[]                                                 # List containing all registered roll numbers
    name={}                                                         # Dictionary for mapping roll numbers to names of students
    date_mapping={}                                                 # Dictionary for mapping each lecture falling on Monday and Thursday to an integer
    
    for i in df.index:                                              # Iterating through the dataframe row-wise
        roll_no=df['Roll No'][i]                                    # Roll no. of student  
        roll_numbers.append(df['Roll No'][i])                       # Appending roll no. of student in the list 'roll_numbers'
        rolls[roll_no]=i                                            # Mapping roll numbers to integers
        name[roll_no]=df['Name'][i]                                 # Mapping roll numbers to names of students
    n=len(rolls)                                                    # Finding number of roll numbers
    
    del df
    try:
        df1=pd.read_csv('input_attendance.csv')                     # Reading input_attendace file 
    except:
        print('Input file error')
    df = df1[df1['Attendance'].notnull()]                           # Removing the rows which have name as empty
    x=len(df)                                                       # No. of rows of dataframe in Attendance file
    format="%d-%m-%Y"                                               # Formatting the date
    start=datetime.strptime(df['Timestamp'][0].split()[0],format)   # Start date of the course           
    end=datetime.strptime(df['Timestamp'][x-1].split()[0],format)   # End date of the course
  
    datesRange = pd.date_range(start, end, freq='D')                # List containing all dates between start date and end date
    lectures=[]                                                     # List for lecture dates
    for i in range(len(datesRange)):
        if(datesRange[i].weekday()==0 or datesRange[i].weekday()==3):       # If dates are on Monday or Thursday append the date in the lectures
            lectures.append(datesRange[i].strftime("%d-%m-%Y"))
    
    for i in range(len(lectures)):                                  # Mapping lecture dates with numbers
        date_mapping[lectures[i]]=i

    matrix=[[[0 for col in range(4)] for col in range(n)] for row in range(len(lectures))]          # 3-d matrix -> lectures*students*(Real+Duplicate+Total+Invalid) Attendance

    for i in df.index:                                              # Iterating through the dataframe of input_attendace file
        try:                                                
            date_and_time=df['Timestamp'][i].split()                # Storing date and time from the column 'Timestamp'
            roll_no_and_name=df['Attendance'][i]                    # Storing roll number and name from the column 'Attendance'
            roll_no=roll_no_and_name.split()[0]                     # Finding roll number by splitting the variable roll_no_and_name
            date=date_and_time[0]                                   # Finding date from date_and_time variable
            time=date_and_time[1].split(':')[0]                     # Finding hour of time from date and time variable by splitting
            
            if(roll_no in roll_numbers and date in lectures):       # Checking if the student roll number and date are valid or not
                matrix[date_mapping[date]][rolls[roll_no]][0]+=1    # Incrementing total attendance by 1
                if(time=='14' or date_and_time[1]=='15:00'):        
                    if(matrix[date_mapping[date]][rolls[roll_no]][1]==0):       # Incrementing real attendance by 1
                        matrix[date_mapping[date]][rolls[roll_no]][1]+=1
                    else:
                        matrix[date_mapping[date]][rolls[roll_no]][2]+=1        # Incrementing duplicate attendance by 1
                else:
                    matrix[date_mapping[date]][rolls[roll_no]][3]+=1            # Incrementing invalid attendance by 1
            
        except:
            continue                                                # If any error occurs move to next row
            
    try:                                                            # Creating output folder if it doesn'y exists
        os.mkdir('output')      
    except:
        pass

    for i in range(n):                                            # For creating individual ROLL_NO.xlsx file
        
        wb=openpyxl.Workbook()
        ws=wb.active
        header_list=['Date','Roll','Name','Total Attendance Count','Real','Duplicate','Invalid','Absent']       # Header of excel file
        for j in range(8):                                                  # Writing the header of the file
            ws.cell(row=1,column=j+1).value=header_list[j]
        ws.cell(row=2,column=2).value=roll_numbers[i]                       # Writing roll number and name
        ws.cell(row=2,column=3).value=name[roll_numbers[i]]
        for j in range(len(lectures)):                                      # Filling the table 
            ws.cell(row=j+3,column=1).value=lectures[j]                     
            for k in range(4):
                ws.cell(row=j+3,column=k+4).value=matrix[j][i][k]
            if(matrix[j][i][1]==0):
                ws.cell(row=j+3,column=8).value=1
            else:
                ws.cell(row=j+3,column=8).value=0
        wb.save(r"output/%s.xlsx"%roll_numbers[i])                          # Saving the individual excel file
    
    wb=openpyxl.Workbook()
    ws=wb.active
    ws.cell(row=1,column=1).value='Roll'                                    # Writing the header of the consolidated file 
    ws.cell(row=1,column=1).value='Name'
    for i in range(len(lectures)):
        ws.cell(row=1,column=2+i).value=lectures[i]
    ws.cell(row=1,column=3+len(lectures)).value='Actual Lecture Taken'
    ws.cell(row=1,column=4+len(lectures)).value='Total Real'
    ws.cell(row=1,column=5+len(lectures)).value='% Attendance'

    for i in range(n):                                                      # Filling the table
        ws.cell(row=i+2,column=1).value=roll_numbers[i]                     # Writing the name and roll number
        ws.cell(row=i+2,column=2).value=name[roll_numbers[i]]
        k=3
        count=0
        for j in range(len(lectures)):                                      # Writing 'P' and 'A'
            if(matrix[j][i][1]==1):
                ws.cell(row=i+2,column=k).value='P'
                count+=1
            else:
                ws.cell(row=i+2,column=k).value='A'
            k+=1
        ws.cell(row=i+2,column=k).value=len(lectures)                       # Writing total lectures, real attendance and percentage attendance
        ws.cell(row=i+2,column=k+1).value=count
        ws.cell(row=i+2,column=k+2).value=round(count*100/len(lectures),2)

    

    wb.save(r"output/attendance_report_consolidated.xlsx")                  # Saving the consolidated file
    def send_email():
        try:
            subject = "Consolidated Attendace Report"
            body = "The report is attached with this mail."
            sender_email = input("Enter sender email : ")
            receiver_email = "sanjuphogat8238@gmail.com"
            password = input("Type your password and press enter:")

            # Create a multipart message and set headers
            message = MIMEMultipart()
            message["From"] = sender_email
            message["To"] = receiver_email
            message["Subject"] = subject
            message["Bcc"] = receiver_email  # Recommended for mass emails

            # Add body to email
            message.attach(MIMEText(body, "plain"))

            filename = "output/attendance_report_consolidated.xlsx"  # In same directory as script

            # Open csv file in binary mode
            with open(filename, "rb") as attachment:
                # Add file as application/octet-stream
                # Email client can usually download this automatically as attachment
                part = MIMEBase("application", "octet-stream")
                part.set_payload(attachment.read())

            # Encode file in ASCII characters to send by email    
            encoders.encode_base64(part)

            # Add header as key/value pair to attachment part
            part.add_header(
                "Content-Disposition",
                f"attachment; filename= {filename}",
            )
            # Add attachment to message and convert message to string
            message.attach(part)
            text = message.as_string()

            # Log in to server using secure context and send email
            context = ssl.create_default_context()
            with smtplib.SMTP_SSL("smtp.gmail.com", 465, context=context) as server:
                server.login(sender_email, password)
                server.sendmail(sender_email, receiver_email, text)
        except:
            print("Error in sending the file.")

    ans=input("Do you want to email to cs3842022@gmail.com (Y/N) : ")
    if(ans=='Y'):
        send_email()
            
        
from platform import python_version
ver = python_version()

if ver == "3.8.10":
    print("Correct Version Installed")
else:
    print("Please install 3.8.10. Instruction are present in the GitHub Repo/Webmail. Url: https://pastebin.com/nvibxmjw")


attendance_report()                                                                                     # Calling the function to do the whole process




#This shall be the last lines of the code.
end_time = datetime.now()
print('Duration of Program Execution: {}'.format(end_time - start_time))