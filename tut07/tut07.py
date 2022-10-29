from asyncore import read										# Importing libraries
from datetime import datetime
from tkinter import SOLID
import openpyxl
from openpyxl import workbook,load_workbook
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from itertools import repeat
from openpyxl.styles.borders import Border, Side

from pandas import read_excel
from pyparsing import col
start_time = datetime.now()
import os


thin_border = Border(left=Side(style='thin'),                   # Defining border
					right=Side(style='thin'), 
					top=Side(style='thin'), 
					bottom=Side(style='thin'))



def find_octant(a,b,c):                                         # Function to find the octant 
	if(a>0 and b>0 and c>0):    
		return 1
	elif(a>0 and b>0 and c<0):
		return -1
	elif(a<0 and b>0 and c>0):
		return 2
	elif(a<0 and b>0 and c<0):
		return -2
	elif(a<0 and b<0 and c>0):
		return 3
	elif(a<0 and b<0 and c<0):
		return -3
	elif(a>0 and b<0 and c>0):
		return 4
	elif(a>0 and b<0 and c<0):
		return -4


def octant_analysis(mod=5000):

	def process_file(s,mod):										# Function to process individual file
		in_file='input/'+s											# Path to file
		df=read_excel(in_file)										# Reading file in dataframe
		wb=Workbook()												# Starting a new workbook
		ws=wb.active													
		lst=['T','U','V','W','U Avg','V Avg','W Avg',r"U'=U-U avg",r"V'=V-V avg",r"W'=W-W avg",'Octant']
		for i in range(11):											# Writing header of file 
			ws.cell(row=2,column=i+1).value=lst[i]	
		octant=[]
		u_avg=df['U'].mean()                                        # Finding average of u,v and w                 
		v_avg=df['V'].mean()
		w_avg=df['W'].mean()
		
		ws.cell(row=1,column=14).value='Overall Octant Count'
		ws.cell(row=1,column=45).value='Longest Subsequence Length'
		ws.cell(row=1,column=49).value='Longest Subsquence Length with Range'

		ws.cell(row=3,column=5).value=round(u_avg,3)
		ws.cell(row=3,column=6).value=round(v_avg,3)
		ws.cell(row=3,column=7).value=round(w_avg,3)

		for i in df.index:
			ws.cell(row=i+3,column=1).value=df['T'][i]
			ws.cell(row=i+3,column=2).value=df['U'][i]
			ws.cell(row=i+3,column=3).value=df['V'][i]
			ws.cell(row=i+3,column=4).value=df['W'][i]
			ws.cell(row=i+3,column=8).value=round(df['U'][i]-u_avg,3)
			ws.cell(row=i+3,column=9).value=round(df['V'][i]-v_avg,3)
			ws.cell(row=i+3,column=10).value=round(df['W'][i]-w_avg,3)
			ws.cell(row=i+3,column=11).value=find_octant(df['U'][i]-u_avg,df['V'][i]-v_avg,df['W'][i]-w_avg)
			octant.append(find_octant(df['U'][i]-u_avg,df['V'][i]-v_avg,df['W'][i]-w_avg))
		
		s=s[:-5]
		file_name='output/'+s+' cm_vel_octant_analysis_mod_'+str(mod)+'.xlsx'
		wb.save(file_name)
	
	os.mkdir('output')
	input_files=os.listdir('input')
	for i in range(len(input_files)):
		process_file(input_files[i],mod)
		


##Read all the excel files in a batch format from the input/ folder. Only xlsx to be allowed
##Save all the excel files in a the output/ folder. Only xlsx to be allowed
## output filename = input_filename[_octant_analysis_mod_5000].xlsx , ie, append _octant_analysis_mod_5000 to the original filename. 

###Code

from platform import python_version
ver = python_version()

if ver == "3.8.10":
	print("Correct Version Installed")
else:
	print("Please install 3.8.10. Instruction are present in the GitHub Repo/Webmail. Url: https://pastebin.com/nvibxmjw")


mod=5000
octant_analysis(mod)






#This shall be the last lines of the code.
end_time = datetime.now()
print('Duration of Program Execution: {}'.format(end_time - start_time))
