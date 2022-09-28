# def octant_transition_count(mod=5000):
# ###Code

# from platform import python_version
# ver = python_version()

# if ver == "3.8.10":
#     print("Correct Version Installed")
# else:
#     print("Please install 3.8.10. Instruction are present in the GitHub Repo/Webmail. Url: https://pastebin.com/nvibxmjw")

# mod=5000
# octant_transition_count(mod)

import pandas as pd
import os
import openpyxl
from openpyxl import workbook,load_workbook

os.chdir(r"C:\Users\Sanju Phogat\Documents\GitHub\2001EE62_2022\tut02")

mod=5000                                              # User input - Mod value

def find_octant(a,b,c):                             # Function to find the octant 
    if(a>0 and b>0 and c>0):    
        return 1
    elif(a>0 and b>0 and c<0):
        return -1
    elif(a<0 and b>0 and c>0):
        return 2
    elif(a<0 and b>0 and c<0):
        return -2
    elif(a<0 and b<0 and c>0):
        return 3
    elif(a<0 and b<0 and c<0):
        return -3
    elif(a>0 and b<0 and c>0):
        return 4
    elif(a>0 and b<0 and c<0):
        return -4

df=pd.read_excel('input_octant_transition_identify.xlsx')
n=len(df['U'])

u_avg=df['U'].mean()                                # Finding average of u,v and w                 
v_avg=df['V'].mean()
w_avg=df['W'].mean()

l1=[u_avg]
l2=[v_avg]
l3=[w_avg]
for i in range(1,n):
    l1.append(" ")
    l2.append(" ")
    l3.append(" ")
df['U Avg']=l1                              # Creating a column of average of u,v and w in dataframe 'df' from dataframe 'x' so that the average columns of u,v and w are of one row
df['V Avg']=l2
df['W Avg']=l3

df[r"U'=U - U Avg"]=round(df['U']-u_avg,5)                   # Subtracting mean of u,v and w from original values and creating a column of it
df[r"V'=V - V Avg"]=round(df['V']-v_avg,5)
df[r"W'=W - W Avg"]=round(df['W']-w_avg,5)

octant=[]                                           # List to store octants

for i in df.index:                                  # Finding octant and storing them in a list named 'Octant'
    octant.append(find_octant(df[r"U'=U - U Avg"][i],df[r"V'=V - V Avg"][i],df[r"W'=W - W Avg"][i]))

df['Octant']=octant                                 # Creating a column for storing corresponding octants in dataframe
df.to_excel('output.xlsx',index=False)

wb=load_workbook('output.xlsx')
ws=wb.active
ws['L3']='User Input'

matrix=[]                                           # 2-d matrix for storing octants within ranges
count=[0]*9                                         # Creating a list for storing elements of 9 columns

count[0]='Octant ID'                                # Storing header list in 'count' list

for i in range(0,4):
    count[2*i+1]=(i+1)
    count[2*(i+1)]=-(i+1)
matrix.append(count)                                # Appending header list in matrix
for i in range(13,22):
    ws.cell(row=1,column=i).value=count[i-13]
count=[0]*9                                         # Resetting values in list 'count'

for i in octant:                                    # Finding total count of values in different octants
    if(i==1):
        count[1]=count[1]+1
    elif(i==-1):
        count[2]=count[2]+1
    elif(i==2):
        count[3]=count[3]+1
    elif(i==-2):
        count[4]=count[4]+1
    elif(i==3):
        count[5]=count[5]+1
    elif(i==-3):
        count[6]=count[6]+1
    elif(i==4):
        count[7]=count[7]+1
    elif(i==-4):
        count[8]=count[8]+1

count[0]='Overall Count'                            
matrix.append(count)                                # Appending total count of values in different octants to the 2-d matrix

matrix.append(['User input','Mod '+str(mod)])         # Appending this value in the matrix

n=len(octant)                                       # Finding the number of points given in the input
count=[0]*9                                         # Resetting the values in the list 'count'
k=0                                                 # Variable to keep track of the index of data we are on
for i in octant:                                    # Counting number of values in different octants in mod range
    if(i==1):
        count[1]=count[1]+1
    elif(i==-1):
        count[2]=count[2]+1
    elif(i==2):
        count[3]=count[3]+1
    elif(i==-2):
        count[4]=count[4]+1
    elif(i==3):
        count[5]=count[5]+1
    elif(i==-3):
        count[6]=count[6]+1
    elif(i==4):
        count[7]=count[7]+1
    elif(i==-4):
        count[8]=count[8]+1
    k=k+1                                           # Incrementing the index tracking variable
    if(k%mod==1):                                     # Processing the mod values in the range and storing them in the list 'count'
        count[0]=str(k-1)+'-'                       
    elif(k%mod==0 or k==n):
        count[0]=count[0]+str(k-1)                  # Here count[0]-> represents the range and further elements of count represents the count in different octants
        matrix.append(count)                        # Appending the list in the matrix
        count=[0]*9                                 # Resetting count of values in different octants



wb.save('output.xlsx')