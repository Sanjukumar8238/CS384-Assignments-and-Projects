import pandas as pd
import os
import openpyxl
from openpyxl import workbook,load_workbook
from openpyxl.styles.borders import Border, Side

os.chdir(r"C:\Users\Sanju Phogat\Documents\GitHub\2001EE62_2022\tut02")

mod=5000                                                        # User input - Mod value

def find_octant(a,b,c):                                         # Function to find the octant 
    if(a>0 and b>0 and c>0):    
        return 1
    elif(a>0 and b>0 and c<0):
        return -1
    elif(a<0 and b>0 and c>0):
        return 2
    elif(a<0 and b>0 and c<0):
        return -2
    elif(a<0 and b<0 and c>0):
        return 3
    elif(a<0 and b<0 and c<0):
        return -3
    elif(a>0 and b<0 and c>0):
        return 4
    elif(a>0 and b<0 and c<0):
        return -4

df=pd.read_excel('input_octant_transition_identify.xlsx')       # Reading input file and storing in dataframe 'df'
n=len(df['U'])                                                  # Finding number of values

u_avg=df['U'].mean()                                            # Finding average of u,v and w                 
v_avg=df['V'].mean()
w_avg=df['W'].mean()

l1=[u_avg]                                                      # Storing values of average of u,v, and w in lists
l2=[v_avg]
l3=[w_avg]
for i in range(1,n):                                            # Appending empty string in all th remaining positions of list
    l1.append(" ")
    l2.append(" ")
    l3.append(" ")
df['U Avg']=l1                                                  # Adding the lists as column in dataframe
df['V Avg']=l2
df['W Avg']=l3

df[r"U'=U - U Avg"]=df['U']-u_avg                               # Subtracting mean of u,v and w from original values and creating a column of it
df[r"V'=V - V Avg"]=df['V']-v_avg
df[r"W'=W - W Avg"]=df['W']-w_avg

octant=[]                                                       # List to store octants

for i in df.index:                                              # Finding octant and storing them in a list named 'Octant'
    octant.append(find_octant(df[r"U'=U - U Avg"][i],df[r"V'=V - V Avg"][i],df[r"W'=W - W Avg"][i]))

df['Octant']=octant                                             # Creating a column for storing corresponding octants in dataframe
df.to_excel('output_octant_transition_identify.xlsx',index=False)                          # Saving the dataframe in file

wb=load_workbook('output_octant_transition_identify.xlsx')                                 # Loading the file in workbook
ws=wb.active
ws['L3']='User Input'                                           # Putting the string 'User Input' at its specified place

matrix=[]                                                       # 2-d matrix for storing octants within ranges
count=[0]*9                                                     # Creating a list for storing elements of 9 columns

count[0]='Octant ID'                                            # Storing header list in 'count' list

thin_border = Border(left=Side(style='thin'),                   # Defining border
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))

for i in range(0,4):
    count[2*i+1]=(i+1)
    count[2*(i+1)]=-(i+1)
matrix.append(count)                                            # Appending header list in matrix
for i in range(13,22):                                          # Writing header list in worksheet
    ws.cell(row=1,column=i).value=count[i-13]
    ws.cell(row=1,column=i).border=thin_border
count=[0]*9                                                     # Resetting values in list 'count'

for i in octant:                                                # Finding total count of values in different octants
    if(i==1):
        count[1]=count[1]+1
    elif(i==-1):
        count[2]=count[2]+1
    elif(i==2):
        count[3]=count[3]+1
    elif(i==-2):
        count[4]=count[4]+1
    elif(i==3):
        count[5]=count[5]+1
    elif(i==-3):
        count[6]=count[6]+1
    elif(i==4):
        count[7]=count[7]+1
    elif(i==-4):
        count[8]=count[8]+1

count[0]='Overall Count'                                        # Creating overall count row
matrix.append(count)                                           
for i in range(13,22):                                          # Writing overall count in worksheet
    ws.cell(row=2,column=i).value=count[i-13]
    ws.cell(row=2,column=i).border=thin_border                           
ws.cell(row=3,column=13).value='Mod '+str(mod)                  # Writing mod value at specified cell
for i in range(13,22):                                         
    ws.cell(row=3,column=i).border=thin_border                           


n=len(octant)                                                   # Finding the number of points given in the input
count=[0]*9                                                     # Resetting the values in the list 'count'
k=0                                                             # Variable to keep track of the index of data we are on
j=4                                                             # Variable to keep track of row in worksheet
for i in octant:                                                # Counting number of values in different octants in mod range
    if(i==1):
        count[1]=count[1]+1
    elif(i==-1):
        count[2]=count[2]+1
    elif(i==2):
        count[3]=count[3]+1
    elif(i==-2):
        count[4]=count[4]+1
    elif(i==3):
        count[5]=count[5]+1
    elif(i==-3):
        count[6]=count[6]+1
    elif(i==4):
        count[7]=count[7]+1
    elif(i==-4):
        count[8]=count[8]+1
    k=k+1                                                       # Incrementing the index tracking variable
    if(k%mod==1):                                               # Processing the mod values in the range and storing them in the list 'count'
        count[0]=str(k-1)+'-'                       
    elif(k%mod==0 or k==n):
        count[0]=count[0]+str(k-1)                              # Here count[0]-> represents the range and further elements of count represents the count in different octants
        for i in range(13,22):                                  # Writing the mod count of octant in worksheet
            ws.cell(row=j,column=i).value=count[i-13]
            ws.cell(row=j,column=i).border=thin_border
        j=j+1                                                   # Incrementing row
        matrix.append(count)
        count=[0]*9                                             # Resetting count of values in different octants

count[0]='Verified'                                             # Storing the verified column in 'count' list
for i in range(1,9):
    sum=0
    for z in range(2,len(matrix)):
        sum+=matrix[z][i]
    count[i]=sum

for i in range(13,22):                                          # Writing verified row in worksheet
    ws.cell(row=j,column=i).value=count[i-13] 
    ws.cell(row=j,column=i).border=thin_border  
j+=3

ws.cell(row=j,column=13).value='Overall Transition Count'       # Writing overall transition count in worksheet
ws.cell(row=j+3,column=12).value='From'
ws.cell(row=j+1,column=14).value='To'
ws.cell(row=j+1,column=13).value=str(0)+'-'+str(n-1)
j+=2

matrix.clear()
matrix = [ [0]*9 for i in range(9)]                             # Creating 9*9 matrix for storing transition count values

for i in range(0,4):                                            # Storing header row and header column in the matrix
    matrix[0][2*i+1]=(i+1)
    matrix[0][2*(i+1)]=-(i+1)
for i in range(0,9):
    matrix[i][0]=matrix[0][i]
matrix[0][0]='Count'

dic={}                                                          # creating dictionary for mapping 
for i in range(0,4):
    dic[i+1]=2*i+1
    dic[-(i+1)]=2*(i+1)

def find_row_col(x,y):                                          # Finding row and column of matrix from transition values
    lst=[dic[x],dic[y]]
    return lst

prev=octant[0]                                              
for i in range(1,n):                                            # Filling overall transition matrix
    lst=find_row_col(prev,octant[i])                            # lst[0]-> row and lst[1]->column of overall transition matrix
    matrix[lst[0]][lst[1]]+=1
    prev=octant[i]

for i in range(0,9):                                            # Writing the overall transition matrix in worksheet
    for k in range(13,22):
        ws.cell(row=j+i,column=k).value=matrix[i][k-13]
        ws.cell(row=j+i,column=k).border=thin_border
        if(i!=0 and k!=13):
            matrix[i][k-13]=0
       

                                        


wb.save('output_octant_transition_identify.xlsx')                                          # Saving the file