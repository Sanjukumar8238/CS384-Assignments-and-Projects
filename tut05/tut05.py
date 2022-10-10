import pandas as pd                                                 # Importing required modules
import os
import openpyxl
from openpyxl import workbook,load_workbook
from openpyxl.styles.borders import Border, Side
from datetime import datetime
start_time = datetime.now()

mod=5000                                                            # User Input

#Help https://youtu.be/N6PBd4XdnEw
def octant_range_names(mod=5000):
    octant_name_id_mapping = {"1":"Internal outward interaction", "-1":"External outward interaction", "2":"External Ejection", "-2":"Internal Ejection", "3":"External inward interaction", "-3":"Internal inward interaction", "4":"Internal sweep", "-4":"External sweep"}
    dic={}                                                          # creating dictionary for mapping 
    opp_dic={}                                                      # Creating dictionary with opposite key value pair than 'dic'
    
    for i in range(0,4):                                            # dic[1]=0,dic[-1]=-1,...
        dic[i+1]=2*i+1-1                                            # opp_dic[0]=1,opp_dic[1]=-1,...
        dic[-(i+1)]=2*(i+1)-1
        opp_dic[2*i+1-1]=i+1
        opp_dic[2*(i+1)-1]=-(i+1)
    
    def find_octant(a,b,c):                                         # Function to find the octant 
        if(a>0 and b>0 and c>0):    
            return 1
        elif(a>0 and b>0 and c<0):
            return -1
        elif(a<0 and b>0 and c>0):
            return 2
        elif(a<0 and b>0 and c<0):
            return -2
        elif(a<0 and b<0 and c>0):
            return 3
        elif(a<0 and b<0 and c<0):
            return -3
        elif(a>0 and b<0 and c>0):
            return 4
        elif(a>0 and b<0 and c<0):
            return -4

    try:
        df=pd.read_excel('octant_input.xlsx')                       # Reading input file and storing in dataframe 'df'
    except:
        print("File opening error")
        exit()

    n=len(df['U'])                                                  # Finding number of values

    try:
        u_avg=df['U'].mean()                                        # Finding average of u,v and w                 
        v_avg=df['V'].mean()
        w_avg=df['W'].mean()
    except:
        print("Error in values of points")
        exit()

    l1=[u_avg]                                                      # Storing values of average of u,v, and w in lists
    l2=[v_avg]
    l3=[w_avg]
    for i in range(1,n):                                            # Appending empty string in all th remaining positions of list
        l1.append(" ")
        l2.append(" ")
        l3.append(" ")
    df['U Avg']=l1                                                  # Adding the lists as column in dataframe
    df['V Avg']=l2
    df['W Avg']=l3

    df[r"U'=U - U Avg"]=df['U']-u_avg                               # Subtracting mean of u,v and w from original values and creating a column of it
    df[r"V'=V - V Avg"]=df['V']-v_avg
    df[r"W'=W - W Avg"]=df['W']-w_avg

    octant=[]                                                       # List to store octants

    for i in df.index:                                              # Finding octant and storing them in a list named 'Octant'
        octant.append(find_octant(df[r"U'=U - U Avg"][i],df[r"V'=V - V Avg"][i],df[r"W'=W - W Avg"][i]))

    df['Octant']=octant                                             # Creating a column for storing corresponding octants in dataframe

    try:
        df.to_excel('octant_output_ranking_excel.xlsx',index=False)                          # Saving the dataframe in file
    except:
        print("Error in writing to output file")
        exit()

    try:
        wb=load_workbook('octant_output_ranking_excel.xlsx')                                 # Loading the file in workbook
    except:
        print("Error in loading output file")
        exit()

    def find_rank_of_list(lst):                                     # Function to find the rank list from count values of all octants
        temp_lst=lst.copy()
        temp_lst.sort(reverse=True)
        res=[]

        for i in lst:
            for j in range(0,8):
                if(i==temp_lst[j]):
                    res.append(j+1)
        return res                                                  # Returning the ranked list
    
    def find_1st_rank(lst):                                         # Finding the octant which has rank 1 in the given rank list
        for i in range(8):
            if(lst[i]==1):
                return opp_dic[i]

    def count_rank1(lst,x):                                         # Finding the count of rank 1 in the rank 1 mod values of octant x
        sum=0
        for i in lst:
            if(x==i):
                sum+=1
        return sum                                                  # Return the count
    
    rank_matrix=[]                                                  # Matrix to store rank list for different mod values
    rank1_list=[]                                                   # List to store the octants which have rank 1 in different mod ranges and overall
    ws=wb.active
    ws['L4']='User Input'                                           # Putting the string 'User Input' at its specified place

    matrix=[]                                                       # 2-d matrix for storing octants within ranges
    count=[0]*9                                                     # Creating a list for storing elements of 9 columns

    count[0]='Octant ID'                                            # Storing header list in 'count' list

    for i in range(0,4):
        count[2*i+1]=(i+1)
        count[2*(i+1)]=-(i+1)
    matrix.append(count)                                            # Appending header list in matrix
    for i in range(13,22):                                          # Writing header list in worksheet
        ws.cell(row=2,column=i).value=count[i-13]
        if(i>13):
            ws.cell(row=1,column=i+8).value=count[i-13]
            ws.cell(row=2,column=i+8).value='Rank '+str(i-13)
    ws.cell(row=2,column=30).value='Rank1 Octant ID'
    ws.cell(row=2,column=31).value='Rank1 Octant Name'
    count=[0]*9                                                     # Resetting values in list 'count'

    for i in octant:                                                # Finding total count of values in different octants
        if(i==1):
            count[1]=count[1]+1
        elif(i==-1):
            count[2]=count[2]+1
        elif(i==2):
            count[3]=count[3]+1
        elif(i==-2):
            count[4]=count[4]+1
        elif(i==3):
            count[5]=count[5]+1
        elif(i==-3):
            count[6]=count[6]+1
        elif(i==4):
            count[7]=count[7]+1
        elif(i==-4):
            count[8]=count[8]+1

    count[0]='Overall Count'                                        # Creating overall count row
    matrix.append(count)                                           
    for i in range(13,22):                                          # Writing overall count in worksheet
        ws.cell(row=3,column=i).value=count[i-13]
    count.pop(0)                                                    # Removing the header from list
    rank=find_rank_of_list(count)                                   # Find the rank list 
    rank1_list.append(find_1st_rank(rank))                          # Finding the rank 1 octant and appending in rank1_list
    rank_matrix.append(rank)                                        # Appending rank list in the matrix
    for i in range(8):                                              # Writing overall count in worksheet
        ws.cell(row=3,column=22+i).value=rank_matrix[0][i]
    ws.cell(row=3,column=30).value=rank1_list[0]
    ws.cell(row=3,column=31).value=octant_name_id_mapping[str(rank1_list[0])]
    ws.cell(row=4,column=13).value='Mod '+str(mod)                  # Writing mod value at specified cell
                            


    n=len(octant)                                                   # Finding the number of points given in the input
    count=[0]*9                                                     # Resetting the values in the list 'count'
    k=0                                                             # Variable to keep track of the index of data we are on
    j=4                                                             # Variable to keep track of row in worksheet
    for i in octant:                                                # Counting number of values in different octants in mod range
        if(i==1):
            count[1]=count[1]+1
        elif(i==-1):
            count[2]=count[2]+1
        elif(i==2):
            count[3]=count[3]+1
        elif(i==-2):
            count[4]=count[4]+1
        elif(i==3):
            count[5]=count[5]+1
        elif(i==-3):
            count[6]=count[6]+1
        elif(i==4):
            count[7]=count[7]+1
        elif(i==-4):
            count[8]=count[8]+1
        k=k+1                                                       # Incrementing the index tracking variable
        if(k%mod==1):                                               # Processing the mod values in the range and storing them in the list 'count'
            count[0]=str(k-1)+'-'                       
        elif(k%mod==0 or k==n):
            count[0]=count[0]+str(k-1)                              # Here count[0]-> represents the range and further elements of count represents the count in different octants
            for i in range(13,22):                                  # Writing the mod count of octant in worksheet
                ws.cell(row=j+1,column=i).value=count[i-13]
            count.pop(0)                                            # Removing the header from list
            rank=find_rank_of_list(count)                           # Find the rank list 
            rank1_list.append(find_1st_rank(rank))                  # Finding the rank 1 octant and appending in rank1_list
            rank_matrix.append(rank)                                # Appending rank list in the matrix

            for i in range(8):                                                  # Writing the columns of rank, rank1 and octant_name in the worksheet
                ws.cell(row=j+1,column=22+i).value=rank_matrix[j-3][i]
            ws.cell(row=j+1,column=30).value=rank1_list[j-3]
            ws.cell(row=j+1,column=31).value=octant_name_id_mapping[str(rank1_list[j-3])]

            j=j+1                                                   # Incrementing row
            matrix.append(count)
            count=[0]*9                                             # Resetting count of values in different octants    
                                            
    rank1_list.pop(0)                                               # Removing the overall rank1 octant

    ws.cell(row=14,column=14).value='Octant ID'                     # Writing the header of table of count of rank1 mod values
    ws.cell(row=14,column=15).value='Octant Name'
    ws.cell(row=14,column=16).value='Count of Rank 1 Mod Values'

    for i in range(8):                                              # Writing the table of count of rank1 mod values
        ws.cell(row=15+i,column=14).value=opp_dic[i]
        ws.cell(row=15+i,column=15).value=octant_name_id_mapping[str(opp_dic[i])]
        ws.cell(row=15+i,column=16).value=count_rank1(rank1_list,opp_dic[i])


    try:
        wb.save('octant_output_ranking_excel.xlsx')                                          # Saving the file
    except:
        print("Error in saving the output file")


from platform import python_version
ver = python_version()

if ver == "3.8.10":
    print("Correct Version Installed")
else:
    print("Please install 3.8.10. Instruction are present in the GitHub Repo/Webmail. Url: https://pastebin.com/nvibxmjw")



octant_range_names(mod)                                                                     # Calling the function to perform the tasks



#This shall be the last lines of the code.
end_time = datetime.now()
print('Duration of Program Execution: {}'.format(end_time - start_time))
