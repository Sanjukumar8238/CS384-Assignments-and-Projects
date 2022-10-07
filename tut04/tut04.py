import pandas as pd
import os
import openpyxl
from openpyxl import workbook,load_workbook
from openpyxl.styles.borders import Border, Side
from itertools import repeat

from datetime import datetime
start_time = datetime.now()

from platform import python_version
ver = python_version()

if ver == "3.8.10":
    print("Correct Version Installed")
else:
    print("Please install 3.8.10. Instruction are present in the GitHub Repo/Webmail. Url: https://pastebin.com/nvibxmjw")

#Help https://youtu.be/H37f_x4wAC0
def octant_longest_subsequence_count_with_range():
    def find_octant(a,b,c):                                         # Function to find the octant 
        if(a>0 and b>0 and c>0):    
            return 1
        elif(a>0 and b>0 and c<0):
            return -1
        elif(a<0 and b>0 and c>0):
            return 2
        elif(a<0 and b>0 and c<0):
            return -2
        elif(a<0 and b<0 and c>0):
            return 3
        elif(a<0 and b<0 and c<0):
            return -3
        elif(a>0 and b<0 and c>0):
            return 4
        elif(a>0 and b<0 and c<0):
            return -4
    try:
        df=pd.read_excel('input_octant_longest_subsequence_with_range.xlsx')       # Reading input file and storing in dataframe 'df'
    except:
        print("File open error")
        exit()

    n=len(df['U'])                                                  # Finding number of values

    try:
        u_avg=df['U'].mean()                                            # Finding average of u,v and w                 
        v_avg=df['V'].mean()
        w_avg=df['W'].mean()
    except:
        print("Error in the values of input")
        exit()

    l1=[u_avg]                                                      # Storing values of average of u,v, and w in lists
    l2=[v_avg]
    l3=[w_avg]
    for i in range(1,n):                                            # Appending empty string in all th remaining positions of list
        l1.append(" ")
        l2.append(" ")
        l3.append(" ")
    df['U Avg']=l1                                                  # Adding the lists as column in dataframe
    df['V Avg']=l2
    df['W Avg']=l3

    df[r"U'=U - U Avg"]=df['U']-u_avg                               # Subtracting mean of u,v and w from original values and creating a column of it
    df[r"V'=V - V Avg"]=df['V']-v_avg
    df[r"W'=W - W Avg"]=df['W']-w_avg

    octant=[]                                                       # List to store octants

    for i in df.index:                                              # Finding octant and storing them in a list named 'Octant'
        octant.append(find_octant(df[r"U'=U - U Avg"][i],df[r"V'=V - V Avg"][i],df[r"W'=W - W Avg"][i]))

    df['Octant']=octant                                             # Creating a column for storing corresponding octants in dataframe
    df.to_excel('output_octant_longest_subsequence_with_range.xlsx',index=False)                          # Saving the dataframe in file

    try:
        wb=load_workbook('output_octant_longest_subsequence_with_range.xlsx')                                 # Loading the file in workbook
    except:
        print("Error in loading previously written file")
        exit()

    ws=wb.active

    thin_border = Border(left=Side(style='thin'),                   # Defining border
                        right=Side(style='thin'), 
                        top=Side(style='thin'), 
                        bottom=Side(style='thin'))

    r=['Count','Longest Subsequence Length','Count']                # Header list
    for i in range(3):                                              # Writing header of table to worksheet
        ws.cell(row=2,column=13+i).value=r[i] 
        ws.cell(row=2,column=13+i).border=thin_border    


    octants=[]
    for i in range(2,10,2):                                         # Writing octants on leftmost column of the table
        ws.cell(row=i+1,column=13).value=i//2
        octants.append(i//2)
        ws.cell(row=i+2,column=13).value=-(i//2) 
        octants.append(-i//2)
        ws.cell(row=i+1,column=13).border=thin_border
        ws.cell(row=i+2,column=13).border=thin_border                                

    dic={}                                                          # creating dictionary for mapping 
    for i in range(0,4):                                            
        dic[i+1]=2*i+1-1
        dic[-(i+1)]=2*(i+1)-1
            
    count=[0]*8                                                     # List for storing number of longest subsequence
    longest_length=[0]*8                                            # List for storing length of longest subsequence
    prev=octant[0]
    l=1                                                             # Length of current octant

    temp=[0]                                                        # Temporary variable to store range
    ranges= [[] for x in repeat(None, 8)]                           # Empty list of list to store ranges for different octants

    for i in range(1,n+1):                                          # Loop for finding number and length of longest subsequence
        if(i==n):                                                   # IF last is reached process the whole
            if(longest_length[dic[prev]]<l):                        
                longest_length[dic[prev]]=l
                count[dic[prev]]=1
                temp.append(df['Time'][i-1])
                ranges[dic[prev]].clear()
                ranges[dic[prev]].append(temp)
            elif(longest_length[dic[prev]]==l):
                count[dic[prev]]+=1
                temp.append(df['Time'][i-1])
                ranges[dic[prev]].append(temp)
        elif(prev==octant[i]):                                      # If prev and current values are same, increase current length by 1
            l+=1
        else:                                                       # Else process the previous octant values and start with new octant
            if(longest_length[dic[prev]]<l):
                longest_length[dic[prev]]=l
                count[dic[prev]]=1
                ranges[dic[prev]].clear()
                temp.append(df['Time'][i-1])
                ranges[dic[prev]].append(temp)
            elif(longest_length[dic[prev]]==l):
                count[dic[prev]]+=1
                temp.append(df['Time'][i-1])
                ranges[dic[prev]].append(temp)
            temp=[df['Time'][i]]
            l=1
            prev=octant[i]

    
    for i in range(2,10):                                           # Writing the number and length of longest subsequence in table
        ws.cell(row=i+1,column=14).value=longest_length[i-2]
        ws.cell(row=i+1,column=15).value=count[i-2]
        ws.cell(row=i+1,column=14).border=thin_border
        ws.cell(row=i+1,column=15).border=thin_border
    k=2
    ws.cell(row=k,column=17).value='Count'
    ws.cell(row=k,column=18).value='Longest Subsequence Length'
    ws.cell(row=k,column=19).value='Count'
    
    k+=1
    for i in range(8):
        ws.cell(row=k,column=17).value=octants[i]
        ws.cell(row=k,column=18).value=longest_length[i]
        ws.cell(row=k,column=19).value=count[i]
        ws.cell(row=k+1,column=17).value='Time'
        ws.cell(row=k+1,column=18).value='From'
        ws.cell(row=k+1,column=19).value='To'
        x=ranges[i]
        k+=2
        for j in x:
            ws.cell(row=k,column=18).value=j[0]
            ws.cell(row=k,column=19).value=j[1]
            k+=1


    try:
        wb.save('output_octant_longest_subsequence_with_range.xlsx')                                          # Saving the file
    except:
        print("Error in saving the output file")


octant_longest_subsequence_count_with_range()




#This shall be the last lines of the code.
end_time = datetime.now()
print('Duration of Program Execution: {}'.format(end_time - start_time))
