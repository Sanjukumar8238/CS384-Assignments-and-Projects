import os
import streamlit as st

from io import BytesIO
from datetime import datetime
from threading import Thread
import pandas as pd #importing pandas
import numpy as np #importing numpy
import itertools    #This is to iterate through multiple lists
import os          ##importing os
import openpyxl   ##importing openpyxl for excel
import operator
from openpyxl.styles.borders import Border, Side  ## importing some styles
from openpyxl.styles import PatternFill
from openpyxl import Workbook,load_workbook  ##from openpyxl taking workbook and load_book
start_time = datetime.now()

year = start_time.strftime("%Y")
month = start_time.strftime("%m")
day = start_time.strftime("%d")
time = start_time.strftime("%H:%M:%S")
time=time.split(':')
#Help

path="input/"
All_input_files=[]
dir_list = os.listdir(path)
for i in dir_list:
    All_input_files.append(i)
choice = st.sidebar.multiselect("Select file", All_input_files)#for multiple select

mod=st.number_input("Please select a MOD value ",min_value=1000)#giving minimum value as 1000
st.write(f"Your mod value is {mod}")


    
# print(choice)



if st.button('compute'):#for computation
        
    for File in choice:
        try:
            wb=load_workbook("input\\"+File)
            df=pd.read_excel("input\\"+File)
            ws=wb.active
        except:
            print('error')
            exit()
        
        x=df['U'].mean()
        x=round(x,3)
        ws['E1']='U Avg'
        ws['E2']=x
        y=df['V'].mean()
        y=round(y,3)
        ws['F1']='V Avg'
        ws['F2']=y
        z=df['W'].mean()
        z=round(z,3)
        ws['G1']='W Avg'
        ws['G2']=z
        ##fun to add a cell
        def adding_cell(ROW,COL,val):
            cell_to_write = ws.cell(row=ROW, column=COL)
            cell_to_write.value =val
        ## 3 empty list and appending each iteam value with its average value
        U_avg=[]
        V_avg=[]
        W_avg=[]
        Time=[]
        for (iteam1,iteam2,iteam3,iteam4) in zip (df['U'],df['V'],df['W'],df['T']):
            U_avg.append(round(float(iteam1)-x,3))
            V_avg.append(round(float(iteam2)-y,3))
            W_avg.append(round(float(iteam3)-z,3))
            Time.append(iteam4)
        y=1#row no
        col=8#column no
        adding_cell(y,col,"U'=U-U Avg")
        adding_cell(y,col+1,"V'=V-V Avg")
        adding_cell(y,col+2,"W'=W-W Avg")
        y+=1
        for j in range(len(U_avg)):
            adding_cell(y,col,U_avg[j])
            adding_cell(y,col+1,V_avg[j])
            adding_cell(y,col+2,W_avg[j])
            y+=1
        y=1
        col+=3
        # print(i[:len(i)-5])
        Octant=[]  # -->creating an empty list 
        ## --->>traversing through U_avg,V_avg,W_avg  simultaneously and storing the values in list Octant
        for (a,b,c) in zip(U_avg,V_avg,W_avg):

            if (a>0 and b>0 and c>0):  # --> this tells that a,b,c is 1st octant
                Octant.append(+1)
                
            if (a>0 and b>0 and c<0): # --> this tells that a,b,c is 2nd octant
                Octant.append(-1)
                
            if (a<0 and b>0 and c>0): # --> this tells that a,b,c is 3rd octant
                Octant.append(+2)
                
            if (a<0 and b>0 and c<0): # --> this tells that a,b,c is 4th octant
                Octant.append(-2)
                
            if (a<0 and b<0 and c>0): # --> this tells that a,b,c is 5th octant
                Octant.append(+3)
                    
            if (a<0 and b<0 and c<0): # --> this tells that a,b,c is 6th octant
                Octant.append(-3)
                
            if (a>0 and b<0 and c>0): # --> this tells that a,b,c is 7th octant
                Octant.append(+4)
                
            if (a>0 and b<0 and c<0): # --> this tells that a,b,c is 8th octant
                Octant.append(-4)
        adding_cell(y,col,"Octant")
        y+=1  
        for j in range(len(Octant)):
            adding_cell(y,col,Octant[j])
            y+=1
        y=1  
        col+=3
            ## this is for making boarder in the cell
        thin_border = Border(left=Side(style='thin'), 
                            right=Side(style='thin'), 
                            top=Side(style='thin'), 
                            bottom=Side(style='thin'))

        def add_cell(ROW,COL,val):
            cell_to_write = ws.cell(row=ROW, column=COL)
            cell_to_write.value =val
            ws.cell(row=ROW, column=COL).border = thin_border
            ##defing a list
        def fun1():
            y=1
            col=14
            octantID=['Octant ID',1,-1,2,-2,3,-3,4,-4]  ##defing a list 

            OverallCount=['OverallCount',Octant.count(1),Octant.count(-1),Octant.count(2),Octant.count(-2),Octant.count(3),Octant.count(-3),Octant.count(4),Octant.count(-4)]
            adding_cell(y,col,'Overall Octant Count')
            y+=2
            x=0
            try:
                ## putting values of counts of 1,-1 and so on in the given range
                for j in range(col,col+9):
                    
                    add_cell(y,j,octantID[x])
                    
                    add_cell(y+1,j,OverallCount[x])
                    x+=1
                
                add_cell(4,col-1,'Mod '+str(mod))
                
            except:
                print("not typecasting the int value to string Or taking the worng range value")
            
            dicName={1:'Internal outward Interaction',-1:'External outward Interaction',2:'Exteral Ejection',-2:'Interal Ejection',3:'External inward Interaction',-3:'Internal inward Interaction',4:'Internal sweep',-4:'External sweep'}
            ID=[1,-1,2,-2,3,-3,4,-4]
            dic={}
            for  p in range(8):
                dic[ID[p]]=OverallCount[p+1]
            sorted_d = dict( sorted(dic.items(), key=operator.itemgetter(1),reverse=True))
            temp = list(sorted_d.items())
            col=23
            for p in ID:
                res = [idx for idx, key in enumerate(temp) if key[0] == p]
                
                add_cell(4,col,res[0]+1)
                if res[0]+1==1:
                    ws.cell(row=4, column=col).fill= PatternFill(start_color='FFD970', end_color='FFD970', fill_type="solid")##to fill color to the cell
                col+=1

            for new_s, new_val in sorted_d.items():
                                add_cell(4,col,new_s)
                                add_cell(4,col+1,dicName[new_s])
                                break
            rankID=['Rank octant 1','Rank octant -1','Rank octant 2','Rank octant -2','Rank octant 3','Rank octant -3','Rank octant 4','Rank octant -4','Rank1 Octant ID','Rank1 Octant Name']
            a=2
            
            l=0
            for p in range(23,33,1):
                add_cell(a+1,p,rankID[l])
                l+=1
            size=len(U_avg)
            start=0
            q=5
            y=5
            countID=[] 
            try:
                while start<size:
                        x=0
                        
                        r=23
                        arr=[]  ##empty list
                        dic={}
                        
                        if start+mod>size:  ## in case when at the last equal division is not possible
                            # creating a arr which has all counts and range
                            arr=[str(start)+"-"+"Last Index",Octant[start:].count(1),Octant[start:].count(-1),Octant[start:].count(2),Octant[start:].count(-2),Octant[start:].count(3),Octant[start:].count(-3),Octant[start:].count(4),Octant[start:].count(-4)]
                            ## dic of each id and its count in the given range
                            dic={1:Octant[start:].count(1),-1:Octant[start:].count(-1),2:Octant[start:].count(2),-2:Octant[start:].count(-2),3:Octant[start:].count(3),-3:Octant[start:].count(-3),4:Octant[start:].count(4),-4:Octant[start:].count(-4)}
                            ##sorted the dic by its value
                            sorted_d = dict( sorted(dic.items(), key=operator.itemgetter(1),reverse=True))
                            temp = list(sorted_d.items())
                            
                                
                            for p in ID:
                                res = [idx for idx, key in enumerate(temp) if key[0] == p]
                                
                                add_cell(q,r,res[0]+1)
                                if res[0]+1==1:
                                        ws.cell(row=q, column=r).fill= PatternFill(start_color='FFD970', end_color='FFD970', fill_type="solid")##to fill color to the cell
                                r+=1
                            for new_s, new_val in sorted_d.items():
                                countID.append(new_s)
                                add_cell(q,r,new_s)
                                add_cell(q,r+1,dicName[new_s])
                                break
                            
                            ##--> inserting  the arr1by row in the file1
                            
                            for j in range(14,23):
                                add_cell(y,j,arr[x])
                                x+=1
                            
                            break ## breaking the while loop
                            
                        else:
                            # creating a arr which has all counts and range
                            arr=[str(start)+"-"+str(start+mod-1),Octant[start:start+mod].count(1),Octant[start:start+mod].count(-1),Octant[start:start+mod].count(2),Octant[start:start+mod].count(-2),Octant[start:start+mod].count(3),Octant[start:start+mod].count(-3),Octant[start:start+mod].count(4),Octant[start:start+mod].count(-4)]
                            ## dic of each id and its count in the given range
                            dic={1:Octant[start:start+mod].count(1),-1:Octant[start:start+mod].count(-1),2:Octant[start:start+mod].count(2),-2:Octant[start:start+mod].count(-2),3:Octant[start:start+mod].count(3),-3:Octant[start:start+mod].count(-3),4:Octant[start:start+mod].count(4),-4:Octant[start:start+mod].count(-4)}
                            ##sorted the dic by its value
                            sorted_d = dict( sorted(dic.items(), key=operator.itemgetter(1),reverse=True))
                            temp = list(sorted_d.items())
                            # for getting the index
                            for p in ID:
                                res = [idx for idx, key in enumerate(temp) if key[0] == p]
                                
                                add_cell(q,r,res[0]+1)
                                if res[0]+1==1:
                                    ws.cell(row=q, column=r).fill= PatternFill(start_color='FFD970', end_color='FFD970', fill_type="solid")##to fill color to the cell
                                r+=1
                            for new_s, new_val in sorted_d.items():
                                countID.append(new_s)
                                add_cell(q,r,new_s)
                                add_cell(q,r+1,dicName[new_s])
                                break
                            ##--> inserting  the arr by row in the file1
                            for j in range(14,23):
                                add_cell(y,j,arr[x])
                                
                                x+=1
                            start=start+mod   ##incrementing the value of i to i+Mod
                            y=y+1
                            q+=1
                        
            except:
                print("not breaking the while loop at end of range Or mistake in range value")

            
            y=y+2
            add_cell(y,30,'OctantID')
            add_cell(y,31,'Octant Name')
            add_cell(y,32,'Count of Rank 1 Mod Values')
            y+=1
            ## for each id appeared how many times
            for p in ID:
                add_cell(y,30,p)
                add_cell(y,31,dicName[p])
                add_cell(y,32,countID.count(p))
                y+=1

        def fun2():
            col=36
            y=1
            size=len(Octant) ## length of U_avg
            
            try: 
                dic={}  ##empty dictionary
                ## adding values in dic 
                for j in range(0,4):
                    dic[j+1]=2*j+1
                    dic[-(j+1)]=2*(j+1)
            except:
                print("worng way of defining the dictionary ")
            try:
                ## 2D arr of 9*9 and initilizing the values with 0
                rows, cols = (9,9) 
                arr = [[0 for k in range(cols)] for j in range(rows)]
            except:
                print("worng way of taking 2D list")
            try:
                ## adding values in the arr and updating its val through iterating in the loop
                for iteam in range(0,size-1):
                    arr[dic[Octant[iteam]]][dic[Octant[iteam+1]]]+=1
                    
            except:
                print("range is out of list")
            try:
                ## 2D arr of 9*9 and initilizing the values with 0
                rows, cols = (9,9) 
                arr = [[0 for k in range(cols)] for j in range(rows)]
            except:
                print("worng way of taking 2D list")
            try:
                ## adding values in the arr and updating its val through iterating in the loop
                for iteam in range(0,size-1):
                    arr[dic[Octant[iteam]]][dic[Octant[iteam+1]]]+=1
                    
            except:
                print("range is out of list")
            try:
                y=1
                col=36
                ## making table for  Overall Transition Count  
                adding_cell(y,col,'Overall Transition Count')
                y=y+2
                
                adding_cell(y,col+1,'To')
                
                adding_cell(y+2,col-1,'From')
                
                y=y+1

                # list
                id=[1,-1,2,-2,3,-3,4,-4]
                arr[0][0]="count"
                ##adding values in the arr
                for j in range(1,9):
                    arr[0][j]=id[j-1]
                    arr[j][0]=id[j-1]

                col=36
                y=3
                for k in range(y+1,y+9):
                            maxi=-1
                            xy=0
                            for j in range (col+1,col+9):
                                add_cell(k,j,arr[k-y][j-col])
                                if(arr[k-y][j-col]>=maxi):
                                    maxi=arr[k-y][j-col]
                                    xy=j
                            ws.cell(row=k, column=xy).fill= PatternFill(start_color='FFD970', end_color='FFD970', fill_type="solid")##to fill color to the cell
                                
                for k in range(y,y+9):
                    for j in range (col,col+9):
                        add_cell(k,j,arr[k-y][j-col])
                        
                        
                ## giving space by increasing the row          
                y=y+9
            except:
                print("range is out of list")
            try:
                a=0 ##initlizing the value a with 0
                ## iterating  in the range till size
                while a<size:
                    y=y+4
                    ## adding values to some cells
                    
                    adding_cell(y,col,'Mod Transition Count')
                    y=y+1
                    
                    ## in case when a+Mod >size
                    if a+mod<size:
                        
                        adding_cell(y,col,str(a)+"-"+str(a+mod-1))
                    ## when a+Mod<size
                    else:
                        
                        adding_cell(y,col,str(a)+"-"+str(size))
                    
                    adding_cell(y,col+1,'To')
                    
                    adding_cell(y+2,col-1,'From')
                    y=y+1
                    ##adding values in a arr[9][9]
                    rows, cols = (9,9)
                    arr = [[0 for k in range(cols)] for j in range(rows)]
                    
                        
                    if a+mod>size:
                        for iteam in range(a,size-1):
                            arr[dic[Octant[iteam]]][dic[Octant[iteam+1]]]+=1
                        id=[1,-1,2,-2,3,-3,4,-4]
                        arr[0][0]="count"
                        for j in range(1,9):
                            arr[0][j]=id[j-1]
                            arr[j][0]=id[j-1]

                        for k in range(y+1,y+9):
                            maxi=-1
                            xy=0
                            for j in range (col+1,col+9):
                                add_cell(k,j,arr[k-y][j-col])
                                if(arr[k-y][j-col]>=maxi):
                                    maxi=arr[k-y][j-col]
                                    xy=j
                            ws.cell(row=k, column=xy).fill= PatternFill(start_color='FFD970', end_color='FFD970', fill_type="solid")##to fill color to the cell
                                
                        for k in range(y,y+9):
                            for j in range (col,col+9): 
                                add_cell(k,j,arr[k-y][j-col])
                        y=y+9
                        break ##coming out of loop
                        
                    
                    else:
                        for iteam in range(a,a+mod):
                            arr[dic[Octant[iteam]]][dic[Octant[iteam+1]]]+=1
                        
                        id=[1,-1,2,-2,3,-3,4,-4]
                        arr[0][0]="count"
                        for j in range(1,9):
                            arr[0][j]=id[j-1]
                            arr[j][0]=id[j-1]


                        for k in range(y,y+9):
                            
                            for j in range (col,col+9):
                                add_cell(k,j,arr[k-y][j-col])
                                
                        for k in range(y+1,y+9):
                            maxi=-1
                            xy=0
                            for j in range (col+1,col+9):
                                add_cell(k,j,arr[k-y][j-col])
                                if(arr[k-y][j-col]>=maxi):
                                    maxi=arr[k-y][j-col]
                                    xy=j
                            ws.cell(row=k, column=xy).fill= PatternFill(start_color='FFD970', end_color='FFD970', fill_type="solid")##to fill color to the cell
                                
                        y=y+9
                        a=a+mod  ##updating the value of a 
            except:
                print("may be range is out of list Or not breaking the while loop(go infinite if not broken)")

        def fun3():
            adding_cell(1,46,"Longest Subsequence Length")
            adding_cell(1,52,"Longest Subsequence Length with Range")
            try:
                ##creating cells for count,Longest Subsequence Length,count
                add_cell(3,46,'Count')
                add_cell(3,47,'Longest Subsequence Length')
                add_cell(3,48,'Count')
                add_cell(3,50,'Count')
                add_cell(3,51,'Longest Subsequence Length')
                add_cell(3,52,'Count')
            except:
                print("cell name might be wrong or at worng place")
            try:

                id=[1,-1,2,-2,3,-3,4,-4]## list of id
                ##inserting the value of id in the cells
                for iD in range(4,12):
                    add_cell(iD,46,id[iD-4])
            except:
                print("error in taking list id Or error in the range value")
            x1=4 # initilizing x1=2 for row
            x2=4 #initilizing x2=2 for row
            ##---->>> for finding the longest subsequence length of every element of id :: the steps are as follows:
            #first when traversing in octant whenever element is equal to current element of id ,we increment the count value
            ##otherwise we make count =0
            # whenever we incounter count is greater tham maxi we update the value of maxi and no of times it occured that is also incremented
            ##or count is equal to maxi value than also we increment the no of time it occured

            try:

                for iD in id: ##loop for traversing in id
                    count=0  # for count of i at every time 
                    maxi=0  ## storing the maximum length of i occured
                    a=0  ## how many times maximum length of i occured
                    
                    for j in range(len(Octant)):  ##traversing in the list octant
                        
                        if (Octant[j] ==iD):  
                            count+=1 
                            
                        else:
                        
                            count=0
                        if (count>maxi):
                                maxi=count
                                a=1
                        elif(count==maxi and Octant[j]==iD):
                            a+=1
                    
                    
                    add_cell(x1,47,maxi)
                    
                    add_cell(x1,48,a)
                    
                    x1+=1
                    
                    add_cell(x2,50,iD)
                    add_cell(x2,51,maxi)
                    add_cell(x2,52,a)
                    x2+=1
                    add_cell(x2,50,"Time")
                    add_cell(x2,51,"From")
                    add_cell(x2,52,"To")
                    x2+=1
                    res=0
                    ## This is for calculating the time range for longest subsequence length
                    for j in range(0,len(Octant),1):
                        
                        if(Octant[j]==iD):
                            res+=1
                        else:
                            res=0
                        if(res==maxi):
                            add_cell(x2,50,"")
                            add_cell(x2,51,Time[j-maxi+1])
                            add_cell(x2,52,Time[j])
                            res=0
                            x2+=1

            except:
                    print("mistake in calculating the longest subsequence length and its count")

        if __name__ == '__main__':
            Thread(target = fun1).start()
            Thread(target = fun2).start()
            Thread(target = fun3).start()
            
        df2=wb.save(filename='output\\'+str(File[:len(File)-5])+"_"+str(mod)+'_'+year+'-'+month+'-'+day+'-'+time[0]+'-'+time[1]+'-'+time[2]+'.xlsx')
        df2=pd.read_excel('output\\'+str(File[:len(File)-5])+"_"+str(mod)+'_'+year+'-'+month+'-'+day+'-'+time[0]+'-'+time[1]+'-'+time[2]+'.xlsx')
        df2.fillna('', inplace=True)#filling the nan values with empty string
        # print(df2)

        st.write(df2)# displaying on the frontend
        
        
        
        # st.title("hello world")

        
        
            
    
  
		
		

    


##Read all the excel files in a batch format from the input/ folder. Only xlsx to be allowed
##Save all the excel files in a the output/ folder. Only xlsx to be allowed
## output filename = input_filename[_octant_analysis_mod_5000].xlsx , ie, append _octant_analysis_mod_5000 to the original filename. 

###Code


from platform import python_version
ver = python_version()

if ver == "3.8.10":
	print("Correct Version Installed")
else:
	print("Please install 3.8.10. Instruction are present in the GitHub Repo/Webmail. Url: https://pastebin.com/nvibxmjw")


# mod=5000
# octant_analysis(mod)






#This shall be the last lines of the code.
end_time = datetime.now()
print('Duration of Program Execution: {}'.format(end_time - start_time))